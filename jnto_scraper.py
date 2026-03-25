#!/usr/bin/env python3
"""
JNTO Tourism Data Scraper
Downloads and parses the latest "Number of Foreign Visitors to Japan by Nationality/Month"
data from the Japan National Tourism Organization.

Source: https://www.jnto.go.jp/statistics/data/visitors-statistics/
Output: output/JNTO_Extracted_Data.csv
Columns: Date, Country, Country_EN, Visitors, YoY_Change
"""

import os
import re
import csv
import logging
import ssl
import urllib.request
from typing import Optional, Dict, List, Tuple
from bs4 import BeautifulSoup

# Create an SSL context that tolerates self-signed/chain certs (JNTO server quirk)
_SSL_CTX = ssl.create_default_context()
_SSL_CTX.check_hostname = False
_SSL_CTX.verify_mode = ssl.CERT_NONE

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl required: pip install openpyxl")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
JNTO_DIR = os.path.join(BASE_DIR, "pdfs")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(JNTO_DIR, exist_ok=True)

JNTO_STATS_URL = "https://www.jnto.go.jp/statistics/data/visitors-statistics/"

# Japanese → English country name mapping (comprehensive)
COUNTRY_MAP = {
    # Totals
    "総数": "Total",
    # Asia
    "アジア計": "Asia Total",
    "韓国": "South Korea",
    "中国": "China",
    "台湾": "Taiwan",
    "香港": "Hong Kong",
    "マカオ": "Macau",
    "タイ": "Thailand",
    "シンガポール": "Singapore",
    "マレーシア": "Malaysia",
    "インドネシア": "Indonesia",
    "フィリピン": "Philippines",
    "ベトナム": "Vietnam",
    "インド": "India",
    "モンゴル": "Mongolia",
    "その他アジア": "Other Asia",
    # Europe
    "欧州計": "Europe Total",
    "ヨーロッパ計": "Europe Total",
    "英国": "UK",
    "フランス": "France",
    "ドイツ": "Germany",
    "イタリア": "Italy",
    "スペイン": "Spain",
    "ロシア": "Russia",
    "オランダ": "Netherlands",
    "オーストリア": "Austria",
    "スイス": "Switzerland",
    "スウェーデン": "Sweden",
    "デンマーク": "Denmark",
    "ノルウェー": "Norway",
    "フィンランド": "Finland",
    "ベルギー": "Belgium",
    "ポルトガル": "Portugal",
    "ポーランド": "Poland",
    "アイルランド": "Ireland",
    "北欧地域": "Nordic Countries",
    "その他ヨーロッパ": "Other Europe",
    "トルコ": "Turkey",
    "イスラエル": "Israel",
    # North America
    "北米計": "North America Total",
    "北アメリカ計": "North America Total",
    "米国": "USA",
    "カナダ": "Canada",
    "メキシコ": "Mexico",
    "その他北アメリカ": "Other North America",
    # South / Latin America
    "中南米": "Latin America",
    "南アメリカ計": "South America Total",
    "ブラジル": "Brazil",
    "その他南アメリカ": "Other South America",
    # Oceania
    "オセアニア計": "Oceania Total",
    "オーストラリア": "Australia",
    "豪州": "Australia",
    "ニュージーランド": "New Zealand",
    "その他オセアニア": "Other Oceania",
    # Middle East / Africa
    "中東地域": "Middle East",
    "GCC6か国": "GCC 6 Countries",
    "アフリカ計": "Africa Total",
    "アフリカ": "Africa",
    # Other / Stateless
    "無国籍・その他": "Stateless/Other",
    "その他": "Other",
}

# Countries to include in dashboard by default (top priority)
PRIORITY_COUNTRIES = {
    "Total", "South Korea", "China", "Taiwan", "Hong Kong",
    "USA", "Thailand", "Australia", "Singapore", "UK",
    "France", "Germany", "India", "Philippines", "Malaysia",
    "Indonesia", "Vietnam", "Canada",
}

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def get_headers() -> dict:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
    }


def find_latest_excel_url() -> Optional[str]:
    """Scrape JNTO stats page and find the latest nationality/month Excel URL."""
    logger.info(f"Fetching JNTO statistics page: {JNTO_STATS_URL}")
    try:
        req = urllib.request.Request(JNTO_STATS_URL, headers=get_headers())
        with urllib.request.urlopen(req, timeout=30, context=_SSL_CTX) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        logger.error(f"Failed to fetch JNTO page: {e}")
        return None

    soup = BeautifulSoup(html, "html.parser")

    # Look for xlsx links matching the nationality/month pattern
    candidates = []
    for tag in soup.find_all("a", href=True):
        href = tag["href"]
        # Match URLs like /statistics/data/_files/YYYYMMDD_XXXX-N.xlsx
        if re.search(r'_files/\d{8}_\d{4}-\d+\.xlsx', href):
            # Prefer files with "-5" or "-4" suffix (nationality/month files)
            if re.search(r'-(4|5)\.xlsx', href):
                full_url = href if href.startswith("http") else "https://www.jnto.go.jp" + href
                # Extract date from filename for sorting
                date_match = re.search(r'(\d{8})_', href)
                date_str = date_match.group(1) if date_match else "00000000"
                candidates.append((date_str, full_url))

    if not candidates:
        # Fallback: any xlsx file
        for tag in soup.find_all("a", href=True):
            href = tag["href"]
            if href.endswith(".xlsx") and "_files/" in href:
                full_url = href if href.startswith("http") else "https://www.jnto.go.jp" + href
                date_match = re.search(r'(\d{8})_', href)
                date_str = date_match.group(1) if date_match else "00000000"
                candidates.append((date_str, full_url))

    if candidates:
        candidates.sort(reverse=True)
        latest_url = candidates[0][1]
        logger.info(f"Found latest JNTO Excel: {latest_url}")
        return latest_url

    logger.warning("Could not find JNTO Excel URL dynamically. Using hardcoded fallback.")
    return "https://www.jnto.go.jp/statistics/data/_files/20260318_1615-5.xlsx"


def download_file(url: str, local_path: str) -> bool:
    """Download a file from URL to local path."""
    # Check if we already have a recent version
    if os.path.exists(local_path):
        mtime = os.path.getmtime(local_path)
        import time
        age_days = (time.time() - mtime) / 86400
        if age_days < 7:
            logger.info(f"Using cached file: {local_path} (age: {age_days:.1f} days)")
            return True

    logger.info(f"Downloading: {url}")
    try:
        req = urllib.request.Request(url, headers=get_headers())
        with urllib.request.urlopen(req, timeout=60, context=_SSL_CTX) as resp:
            data = resp.read()
        with open(local_path, "wb") as f:
            f.write(data)
        logger.info(f"Downloaded {len(data):,} bytes → {local_path}")
        return True
    except Exception as e:
        logger.error(f"Download failed: {e}")
        return False


def parse_jnto_excel(excel_path: str, start_year: int = 2019) -> List[dict]:
    """
    Parse JNTO Excel workbook.

    Structure per sheet (one sheet = one year):
    - Row 1: Title
    - Row 4: Column headers — alternating month name + '伸率' (YoY%)
    - Row 5+: Data rows — Col A = country (JP), Col B = sub-region, Col C+ = data
    """
    logger.info(f"Parsing Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    records = []

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        if year < start_year:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 5:
            continue

        # Find the header row (contains '1月' or 'January')
        header_row_idx = None
        for i, row in enumerate(rows[:10]):
            row_vals = [str(v) if v is not None else "" for v in row]
            if any("1月" in v or "January" in v for v in row_vals):
                header_row_idx = i
                break

        if header_row_idx is None:
            logger.warning(f"Sheet {year}: could not find header row, skipping")
            continue

        # Parse month column positions from header row
        # Format: col 0=Country, col 1=sub-region, col 2=Jan, col 3=Jan YoY, col 4=Feb...
        header = rows[header_row_idx]
        month_cols: Dict[int, Tuple[int, int]] = {}  # month_num -> (visitor_col, yoy_col)
        month_idx = 0
        for col_idx, val in enumerate(header):
            if col_idx < 2:
                continue  # Skip country/sub-region cols
            val_str = str(val) if val is not None else ""
            if "月" in val_str or (isinstance(val, int) and 1 <= val <= 12):
                # This is a month column
                month_num = month_idx + 1
                visitor_col = col_idx
                yoy_col = col_idx + 1  # Next column is 伸率
                month_cols[month_num] = (visitor_col, yoy_col)
                month_idx += 1
                if month_idx >= 12:
                    break

        if not month_cols:
            # Fallback: assume cols 2,3=Jan; 4,5=Feb; ... (2 cols per month)
            for month_num in range(1, 13):
                visitor_col = 2 + (month_num - 1) * 2
                yoy_col = visitor_col + 1
                month_cols[month_num] = (visitor_col, yoy_col)

        # Parse data rows (start after header row)
        for row in rows[header_row_idx + 1:]:
            if not row or all(v is None for v in row):
                continue

            # Country name: prefer col A, fall back to col B (sub-region)
            country_jp = None
            if row[0] is not None and str(row[0]).strip():
                country_jp = str(row[0]).strip()
            elif len(row) > 1 and row[1] is not None and str(row[1]).strip():
                country_jp = str(row[1]).strip().lstrip("　 ")  # Strip full-width spaces

            if not country_jp:
                continue

            # Map to English
            country_en = COUNTRY_MAP.get(country_jp)
            if country_en is None:
                # Try stripping leading spaces
                country_jp_clean = country_jp.strip()
                country_en = COUNTRY_MAP.get(country_jp_clean, country_jp_clean)

            # Extract monthly data
            for month_num, (visitor_col, yoy_col) in month_cols.items():
                visitors = None
                yoy_change = None

                if visitor_col < len(row) and row[visitor_col] is not None:
                    try:
                        visitors = int(float(row[visitor_col]))
                    except (ValueError, TypeError):
                        pass

                if yoy_col < len(row) and row[yoy_col] is not None:
                    try:
                        yoy_change = round(float(row[yoy_col]), 2)
                    except (ValueError, TypeError):
                        pass

                if visitors is None and yoy_change is None:
                    continue  # Skip months with no data

                date_str = f"{year}/{month_num:02d}"
                records.append({
                    "Date": date_str,
                    "Country_JP": country_jp,
                    "Country": country_en,
                    "Visitors": visitors if visitors is not None else "",
                    "YoY_Change": yoy_change if yoy_change is not None else "",
                })

    wb.close()
    logger.info(f"Parsed {len(records)} JNTO data points from {excel_path}")
    return records


def save_jnto_csv(records: List[dict], csv_path: str):
    """Save JNTO records to CSV."""
    if not records:
        logger.warning("No JNTO records to save.")
        return

    fieldnames = ["Date", "Country_JP", "Country", "Visitors", "YoY_Change"]
    records_sorted = sorted(records, key=lambda r: (r["Date"], r["Country"]))

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records_sorted)

    logger.info(f"Saved {len(records)} JNTO records to {csv_path}")


def run_jnto_pipeline(force_download: bool = False) -> str:
    """
    Full JNTO pipeline: find → download → parse → save CSV.
    Returns path to output CSV.
    """
    output_csv = os.path.join(OUTPUT_DIR, "JNTO_Extracted_Data.csv")
    excel_path = os.path.join(JNTO_DIR, "jnto_latest.xlsx")

    # Step 1: Find and download Excel
    excel_url = find_latest_excel_url()
    if excel_url is None:
        excel_url = "https://www.jnto.go.jp/statistics/data/_files/20260318_1615-5.xlsx"

    if force_download and os.path.exists(excel_path):
        os.remove(excel_path)

    if not download_file(excel_url, excel_path):
        logger.error("Could not download JNTO Excel file.")
        return ""

    # Step 2: Parse
    records = parse_jnto_excel(excel_path, start_year=2003)

    if not records:
        logger.error("No records parsed from JNTO Excel.")
        return ""

    # Step 3: Save
    save_jnto_csv(records, output_csv)

    # Step 4: Validate
    total_records = len(records)
    countries = set(r["Country"] for r in records)
    dates = sorted(set(r["Date"] for r in records))
    logger.info(f"JNTO Summary: {total_records} records, {len(countries)} countries, "
                f"date range {dates[0] if dates else 'N/A'} – {dates[-1] if dates else 'N/A'}")

    # Spot-check Total visitors for a known month
    for r in records:
        if r["Date"] == "2024/12" and r["Country"] == "Total":
            logger.info(f"Spot check 2024/12 Total: {r['Visitors']:,} visitors "
                        f"(YoY: {r['YoY_Change']}%)")
            break

    return output_csv


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="JNTO Tourism Data Scraper")
    parser.add_argument("--force", action="store_true", help="Force re-download")
    args = parser.parse_args()

    out = run_jnto_pipeline(force_download=args.force)
    if out:
        print(f"\nJNTO data saved to: {out}")
    else:
        print("JNTO pipeline failed.")
        exit(1)
